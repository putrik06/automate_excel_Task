import openpyxl

workbook =  openpyxl.load_workbook("Employees.xlsx")
# Print the workbook properties
# print(workbook.properties)

# Explore the workbook sheetnames
print(workbook.sheetnames) # have 3 sheets ['EmployeeData', 'Salaries', 'Skills']

# which workbook is the active one?
print(workbook.active) # commonly the first is the active one.
# >>> <Worksheet "EmployeeData">

# Reference the workbook to perform other operation
sheet = workbook['EmployeeData']

# Getting general information about a sheet
print(sheet.title) # EmployeeData

# getting the active sheet cell
print(sheet.active_cell)

# get the dimension of the sheet
print(sheet.dimensions)

# getting the number of row inside sheet
print(sheet.max_row)

# getting the number of column inside sheet
print(sheet.max_column)

# print the value inside the sheet
for i in sheet.values:
    print(i) # in the form of tuples

# Make a new sheet
workbook.create_sheet('NewSheet')
# save the new worksheet
workbook.save("Employees.xlsx")
# check if the new worksheet is present
print(workbook.sheetnames)

# delete the worksheet
sheet = workbook['NewSheet']
workbook.remove(sheet)

# check if it removed
print(workbook.sheetnames)
workbook.save("Employees.xlsx")

# close the workbook
workbook.close()